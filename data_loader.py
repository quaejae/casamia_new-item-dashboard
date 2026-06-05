# -*- coding: utf-8 -*-
"""raw data 로더.

1단계: 로컬 xlsx(openpyxl)에서 담당자별 시트를 읽어 RawProduct 리스트로 표준화.
2단계(예정): 동일한 load() 시그니처로 Google Sheets(gspread) 구현을 교체.

핵심 설계:
  - 컬럼은 헤더 텍스트(HEADER_ROW)로 탐지하므로 열 순서/추가에 강함.
  - 타임라인 13개월 블록은 MONTH_ROW의 숫자 셀 연속 구간으로 자동 인식.
  - 셀에 삽입된 이미지는 openpyxl 의 이미지 앵커(행)로 제품 행에 매핑.
"""
from __future__ import annotations

import os
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

import config
from schema import RawProduct

# raw data 헤더명 → RawProduct 필드 매핑 (헤더 텍스트 정규화 후 비교)
HEADER_ALIASES = {
    "no.": "no", "no": "no",
    "카테고리": "category",
    "브랜드": "brand",
    "프로젝트명": "project",
    "시리즈명": "series",
    "담당자": "owner",
    "오프/온라인": "channel",
    "사이즈 수": "size_count", "사이즈수": "size_count",
    "컬러 수": "color_count", "컬러수": "color_count",
    "sku 수": "sku_count", "sku수": "sku_count",
    "목표매출": "target_revenue",
    "업체": "vendor",
    "업체 국가": "vendor_country", "업체국가": "vendor_country",
    "주요소재": "material",
    "목표 판매가": "target_price_raw", "목표판매가": "target_price_raw",
    "비고": "note",
    "이미지": "image",
}


def _norm(v) -> str:
    return str(v).strip().lower().replace("\n", "").replace(" ", "") if v is not None else ""


def _txt(v) -> str:
    """셀 값을 표시용 문자열로. None→'', float 정수는 정수로."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _detect_columns(ws) -> Tuple[Dict[str, int], List[int]]:
    """헤더 행을 읽어 {필드명: 열index} 와 타임라인 열 리스트를 반환."""
    field_to_col: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=config.HEADER_ROW, column=col).value
        key = _norm(raw)
        if key in HEADER_ALIASES:
            field_to_col[HEADER_ALIASES[key]] = col

    # 타임라인: MONTH_ROW 에서 숫자 셀이 연속된 구간
    timeline_cols: List[int] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=config.MONTH_ROW, column=col).value
        if isinstance(v, (int, float)):
            timeline_cols.append(col)
    return field_to_col, timeline_cols


def _extract_images(ws) -> Dict[int, Tuple[bytes, str]]:
    """시트의 임베드 이미지를 {앵커 행(1-base): (bytes, ext)} 로 추출."""
    result: Dict[int, Tuple[bytes, str]] = {}
    for img in getattr(ws, "_images", []) or []:
        try:
            anchor = img.anchor
            # OneCellAnchor / TwoCellAnchor 모두 _from.row 보유 (0-base)
            row0 = anchor._from.row  # type: ignore[attr-defined]
        except Exception:
            continue
        row1 = row0 + 1
        try:
            data = img._data()  # openpyxl: 이미지 바이트 반환
        except Exception:
            try:
                data = img.ref.getvalue()  # type: ignore[attr-defined]
            except Exception:
                continue
        ext = (getattr(img, "format", None) or "png").lower()
        result[row1] = (data, ext)
    return result


def _load_sheet(ws) -> List[RawProduct]:
    field_to_col, timeline_cols = _detect_columns(ws)
    images = _extract_images(ws)

    def cell(row: int, field: str):
        col = field_to_col.get(field)
        return ws.cell(row=row, column=col).value if col else None

    def num(row: int, field: str) -> Optional[float]:
        v = cell(row, field)
        return float(v) if isinstance(v, (int, float)) else None

    products: List[RawProduct] = []
    for row in range(config.DATA_START_ROW, ws.max_row + 1):
        # 핵심 식별값(프로젝트명/카테고리)이 모두 비면 빈 행으로 보고 스킵
        project = _txt(cell(row, "project"))
        category = _txt(cell(row, "category"))
        if not project and not category:
            continue

        timeline = [_txt(ws.cell(row=row, column=c).value) for c in timeline_cols]

        img = images.get(row)
        no_val = num(row, "no")

        products.append(RawProduct(
            no=int(no_val) if no_val is not None else None,
            category=category,
            brand=_txt(cell(row, "brand")),
            project=project,
            series=_txt(cell(row, "series")),
            owner=_txt(cell(row, "owner")),
            channel=_txt(cell(row, "channel")),
            size_count=num(row, "size_count"),
            color_count=num(row, "color_count"),
            sku_count=num(row, "sku_count"),
            target_revenue=_txt(cell(row, "target_revenue")),
            vendor=_txt(cell(row, "vendor")),
            vendor_country=_txt(cell(row, "vendor_country")),
            material=_txt(cell(row, "material")),
            target_price_raw=_txt(cell(row, "target_price_raw")),
            timeline=timeline,
            note=_txt(cell(row, "note")),
            image=img[0] if img else None,
            image_ext=img[1] if img else None,
        ))
    return products


def _load_all(wb) -> Dict[str, List[RawProduct]]:
    """워크북의 모든 시트를 읽어 {담당자: [RawProduct]} 반환(시트 순서 보존)."""
    result: Dict[str, List[RawProduct]] = {}
    for ws in wb.worksheets:
        products = _load_sheet(ws)
        if products:
            owner = products[0].owner or ws.title   # owner 우선, 없으면 시트명
            result[owner] = products
    return result


def load(path: Optional[str] = None) -> Dict[str, List[RawProduct]]:
    """로컬 xlsx 를 읽어 {담당자(시트명): [RawProduct, ...]} 반환."""
    path = path or config.LOCAL_XLSX
    if not os.path.isabs(path):
        path = os.path.join(os.getcwd(), path)
    return _load_all(openpyxl.load_workbook(path, data_only=True))


def load_bytes(data: bytes) -> Dict[str, List[RawProduct]]:
    """xlsx 바이트(예: Google Sheets export)를 읽어 동일 형식으로 반환."""
    import io
    return _load_all(openpyxl.load_workbook(io.BytesIO(data), data_only=True))


if __name__ == "__main__":
    import sys
    io_path = sys.argv[1] if len(sys.argv) > 1 else None
    data = load(io_path)
    for owner, prods in data.items():
        imgs = sum(1 for p in prods if p.image)
        print(f"[{owner}] 제품 {len(prods)}개, 이미지 {imgs}개")
        for p in prods[:2]:
            print(f"   - No{p.no} {p.series}/{p.project} | {p.target_price_raw} | "
                  f"{p.vendor}({p.vendor_country}) | tl={sum(1 for t in p.timeline if t)}칸")
